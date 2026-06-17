from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponse
from .models import Utilisateur, Etudiant, Livre, Emprunt, Rapport
from .exceptions import (
    LoginIncorrectException, CompteBloquéException, CompteInactifException,
    LivreIndisponibleException, LivreIntrouvableException,
    MembreIntrouvableException, EmpruntIntrouvableException,
    EmpruntEnCoursException, DoublonException,
    FichierExcelInvalideException, MotDePasseInvalideException
)
from datetime import date, datetime
from io import BytesIO
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4


# ─── TENTATIVES LOGIN ─────────────────────────────────────────
tentatives = {}


# ══════════════════════════════════════════════════════════════
#  LOGIN / LOGOUT
# ══════════════════════════════════════════════════════════════
def login_view(request):
    remembered_username = request.COOKIES.get('remembered_username', '')

    if request.method == 'POST':
        login    = request.POST.get('login', '').strip()
        password = request.POST.get('password', '').strip()
        remember_me = request.POST.get('remember_me') == 'on'

        try:
            if tentatives.get(login, 0) >= 3:
                raise CompteBloquéException()

            try:
                user = Utilisateur.objects.get(login=login)
            except Utilisateur.DoesNotExist:
                raise LoginIncorrectException()

            if not user.actif:
                raise CompteInactifException()

            if user.mot_de_passe != password:
                tentatives[login] = tentatives.get(login, 0) + 1
                raise LoginIncorrectException()

            # Connexion réussie
            tentatives[login] = 0
            request.session['user_id'] = user.id
            request.session['role']    = user.role
            request.session['login']   = user.login

            if remember_me:
                request.session.set_expiry(1209600)  # 2 semaines
            else:
                request.session.set_expiry(0)  # Expire à la fermeture du navigateur

            if user.role == 'etudiant':
                response = redirect('/recherche/')
            elif user.role == 'bibliothecaire':
                response = redirect('/gestion-livres/')
            else:
                response = redirect('/dashboard/')

            if remember_me:
                response.set_cookie('remembered_username', login, max_age=1209600)
            else:
                response.delete_cookie('remembered_username')

            return response

        except CompteBloquéException:
            messages.error(request, "Accès bloqué après 3 tentatives. Contactez l'administrateur.")
        except CompteInactifException:
            messages.error(request, "Votre compte est désactivé. Contactez l'administrateur.")
        except LoginIncorrectException:
            restantes = 3 - tentatives.get(login, 0)
            if restantes > 0:
                messages.error(request, f"Login ou mot de passe incorrect. {restantes} tentative(s) restante(s).")
            else:
                messages.error(request, "Accès bloqué après 3 tentatives.")

    total_livres = Livre.objects.count()
    membres_actifs = Etudiant.objects.count()
    emprunts_en_cours = Emprunt.objects.filter(rendu=False).count()
    livres_disponibles = Livre.objects.filter(disponible=True).count()

    return render(request, 'login.html', {
        'total_livres': total_livres,
        'membres_actifs': membres_actifs,
        'emprunts_en_cours': emprunts_en_cours,
        'livres_disponibles': livres_disponibles,
        'remembered_username': remembered_username,
    })


def logout_view(request):
    request.session.flush()
    return redirect('/login/')


# ══════════════════════════════════════════════════════════════
#  CHANGER MOT DE PASSE (Étudiant)
# ══════════════════════════════════════════════════════════════
def changer_mot_de_passe(request):
    if 'user_id' not in request.session:
        return redirect('/login/')

    if request.method == 'POST':
        ancien     = request.POST.get('ancien_mdp', '').strip()
        nouveau    = request.POST.get('nouveau_mdp', '').strip()
        confirmer  = request.POST.get('confirmer_mdp', '').strip()

        try:
            user = Utilisateur.objects.get(id=request.session['user_id'])

            if user.mot_de_passe != ancien:
                raise MotDePasseInvalideException()

            if len(nouveau) < 4:
                messages.error(request, "Le mot de passe doit contenir au moins 4 caractères.")
            elif nouveau != confirmer:
                messages.error(request, "Les deux mots de passe ne correspondent pas.")
            else:
                user.mot_de_passe = nouveau
                user.save()
                messages.success(request, "Mot de passe changé avec succès !")

        except Utilisateur.DoesNotExist:
            messages.error(request, "Utilisateur introuvable.")
        except MotDePasseInvalideException:
            messages.error(request, "Ancien mot de passe incorrect.")

    return render(request, 'changer_mot_de_passe.html')


# ══════════════════════════════════════════════════════════════
#  RECHERCHE LIVRES (Étudiant)
# ══════════════════════════════════════════════════════════════
def recherche_livre(request):
    if 'user_id' not in request.session:
        return redirect('/login/')

    livres = Livre.objects.all()
    query  = request.GET.get('q', '').strip()
    cat    = request.GET.get('cat', '').strip()
    dispo_only = request.GET.get('dispo', '').strip() == '1'

    if query:
        livres = livres.filter(titre__icontains=query) | \
                 livres.filter(auteur__icontains=query)
    if cat:
        livres = livres.filter(categorie__icontains=cat)
    if dispo_only:
        livres = livres.filter(disponible=True)

    categories = Livre.objects.values_list('categorie', flat=True)\
                              .distinct().exclude(categorie='')

    return render(request, 'recherche.html', {
        'livres':     livres,
        'query':      query,
        'cat':        cat,
        'dispo_only': dispo_only,
        'categories': categories,
    })


# ══════════════════════════════════════════════════════════════
#  MES EMPRUNTS (Étudiant)
# ══════════════════════════════════════════════════════════════
def mes_emprunts(request):
    if 'user_id' not in request.session:
        return redirect('/login/')

    aujourd_hui = date.today()
    try:
        etudiant = Etudiant.objects.get(
            utilisateur__id=request.session['user_id']
        )
        emprunts = Emprunt.objects.filter(etudiant=etudiant)\
                                  .order_by('-date_emprunt')
        en_cours   = emprunts.filter(rendu=False)
        historique = emprunts.filter(rendu=True)
        en_retard  = emprunts.filter(rendu=False, date_retour_prevue__lt=aujourd_hui)
    except Etudiant.DoesNotExist:
        emprunts   = []
        en_cours   = []
        historique = []
        en_retard  = []

    return render(request, 'mes_emprunts.html', {
        'emprunts': emprunts,
        'en_cours': en_cours,
        'historique': historique,
        'en_retard': en_retard,
    })


# ══════════════════════════════════════════════════════════════
#  FICHE ÉTUDIANT
# ══════════════════════════════════════════════════════════════
def fiche_etudiant(request):
    if 'user_id' not in request.session:
        return redirect('/login/')

    aujourd_hui = date.today()
    try:
        etudiant = Etudiant.objects.get(
            utilisateur__id=request.session['user_id']
        )
        emprunts  = Emprunt.objects.filter(etudiant=etudiant).order_by('-date_emprunt')
        en_cours  = emprunts.filter(rendu=False)
        en_retard = emprunts.filter(rendu=False, date_retour_prevue__lt=aujourd_hui)
    except Etudiant.DoesNotExist:
        etudiant  = None
        emprunts  = []
        en_cours  = []
        en_retard = []

    return render(request, 'fiche_etudiant.html', {
        'etudiant': etudiant,
        'emprunts': emprunts,
        'en_cours': en_cours,
        'en_retard': en_retard,
    })


# ══════════════════════════════════════════════════════════════
#  DASHBOARD (Administrateur)
# ══════════════════════════════════════════════════════════════
def dashboard(request):
    if 'user_id' not in request.session:
        return redirect('/login/')

    aujourd_hui        = date.today()
    total_livres       = Livre.objects.count()
    livres_disponibles = Livre.objects.filter(disponible=True).count()
    emprunts_en_cours  = Emprunt.objects.filter(rendu=False).count()
    emprunts_en_retard = Emprunt.objects.filter(
        rendu=False,
        date_retour_prevue__lt=aujourd_hui
    ).count()
    membres_actifs     = Etudiant.objects.count()
    derniers_emprunts  = Emprunt.objects.order_by('-date_emprunt')[:5]

    # Statistiques supplémentaires
    emprunts_rendus    = Emprunt.objects.filter(rendu=True).count()
    comptes_inactifs   = Utilisateur.objects.filter(actif=False).count()
    livres_empruntes   = Livre.objects.filter(disponible=False).count()
    liste_retards      = Emprunt.objects.filter(rendu=False, date_retour_prevue__lt=aujourd_hui)

    # Statistiques par catégorie pour graphique
    from django.db.models import Count
    stats_categories = list(
        Livre.objects.values('categorie')
                     .annotate(total=Count('id'))
                     .order_by('-total')[:5]
    )

    categories_labels  = [sc['categorie'] if sc['categorie'] else "Sans catégorie" for sc in stats_categories]
    categories_valeurs = [sc['total'] for sc in stats_categories]

    # Emprunts par mois (6 derniers mois)
    mois_list = []
    annee, mois = aujourd_hui.year, aujourd_hui.month
    for _ in range(6):
        mois_list.append((annee, mois))
        mois -= 1
        if mois == 0:
            mois = 12
            annee -= 1
    mois_list.reverse()

    nom_mois = {
        1: "Jan", 2: "Fév", 3: "Mar", 4: "Avr", 5: "Mai", 6: "Juin",
        7: "Juil", 8: "Août", 9: "Sept", 10: "Oct", 11: "Nov", 12: "Déc"
    }
    mois_labels  = [f"{nom_mois[m]} {str(y)[2:]}" for y, m in mois_list]
    mois_valeurs = []
    for y, m in mois_list:
        count = Emprunt.objects.filter(date_emprunt__year=y, date_emprunt__month=m).count()
        mois_valeurs.append(count)

    return render(request, 'dashboard.html', {
        'data': {
            'total_livres':       total_livres,
            'livres_disponibles': livres_disponibles,
            'emprunts_en_cours':  emprunts_en_cours,
            'emprunts_en_retard': emprunts_en_retard,
            'membres_actifs':     membres_actifs,
            'emprunts_rendus':    emprunts_rendus,
            'comptes_inactifs':   comptes_inactifs,
            'livres_empruntes':   livres_empruntes,
        },
        'derniers_emprunts':  derniers_emprunts,
        'liste_retards':      liste_retards,
        'categories_labels':  categories_labels,
        'categories_valeurs': categories_valeurs,
        'mois_labels':        mois_labels,
        'mois_valeurs':       mois_valeurs,
    })


# ══════════════════════════════════════════════════════════════
#  GESTION LIVRES (Bibliothécaire)
# ══════════════════════════════════════════════════════════════
def gestion_livres(request):
    if 'user_id' not in request.session:
        return redirect('/login/')

    if request.method == 'POST':
        titre     = request.POST.get('titre', '').strip()
        auteur    = request.POST.get('auteur', '').strip()
        categorie = request.POST.get('categorie', '').strip()
        isbn      = request.POST.get('isbn', '').strip()

        try:
            if not titre or not auteur:
                raise ValueError("Le titre et l'auteur sont obligatoires.")
            if Livre.objects.filter(titre__iexact=titre).exists():
                raise DoublonException(f"Le livre '{titre}' existe déjà.")

            Livre.objects.create(
                titre=titre, auteur=auteur,
                categorie=categorie, isbn=isbn
            )
            messages.success(request, f"Livre '{titre}' ajouté avec succès !")

        except DoublonException as e:
            messages.error(request, str(e))
        except ValueError as e:
            messages.error(request, str(e))

    # Suppression
    suppr_id = request.GET.get('suppr')
    if suppr_id:
        try:
            livre = Livre.objects.get(id=suppr_id)
            if not livre.disponible:
                raise LivreIndisponibleException(
                    f"Impossible — '{livre.titre}' est actuellement emprunté."
                )
            livre.delete()
            messages.success(request, "Livre supprimé.")
        except Livre.DoesNotExist:
            messages.error(request, "Livre introuvable.")
        except LivreIndisponibleException as e:
            messages.error(request, str(e))

    # Filtrage et recherche
    livres     = Livre.objects.all()
    query      = request.GET.get('q', '').strip()
    cat        = request.GET.get('cat', '').strip()
    dispo_only = request.GET.get('dispo', '').strip()

    if query:
        livres = livres.filter(titre__icontains=query) | \
                 livres.filter(auteur__icontains=query)
    if cat:
        livres = livres.filter(categorie__icontains=cat)
    if dispo_only == '1':
        livres = livres.filter(disponible=True)
    elif dispo_only == '0':
        livres = livres.filter(disponible=False)

    categories = Livre.objects.values_list('categorie', flat=True)\
                              .distinct().exclude(categorie='')

    return render(request, 'gestion_livres.html', {
        'livres':     livres.order_by('titre'),
        'query':      query,
        'cat':        cat,
        'dispo_only': dispo_only,
        'categories': categories,
    })


# ── Import Excel Livres ────────────────────────────────────────
def import_livres_excel(request):
    if 'user_id' not in request.session:
        return redirect('/login/')

    if request.method == 'POST':
        fichier = request.FILES.get('fichier_excel')

        try:
            if not fichier:
                raise FichierExcelInvalideException("Aucun fichier sélectionné.")
            if not fichier.name.endswith(('.xlsx', '.xls')):
                raise FichierExcelInvalideException("Le fichier doit être au format Excel (.xlsx).")

            wb      = openpyxl.load_workbook(fichier)
            ws      = wb.active
            headers = [str(cell.value).strip().lower() if cell.value else ''
                       for cell in ws[1]]

            colonnes_requises = ['titre', 'auteur']
            for col in colonnes_requises:
                if col not in headers:
                    raise FichierExcelInvalideException(
                        f"Colonne '{col}' manquante. Colonnes requises : titre, auteur, isbn, categorie"
                    )

            ajoutes  = 0
            ignores  = 0
            erreurs  = []

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    data = dict(zip(headers, row))
                    titre  = str(data.get('titre', '') or '').strip()
                    auteur = str(data.get('auteur', '') or '').strip()

                    if not titre or not auteur:
                        ignores += 1
                        continue

                    if Livre.objects.filter(titre__iexact=titre).exists():
                        ignores += 1
                        continue

                    Livre.objects.create(
                        titre     = titre,
                        auteur    = auteur,
                        isbn      = str(data.get('isbn', '') or '').strip(),
                        categorie = str(data.get('categorie', '') or '').strip(),
                    )
                    ajoutes += 1

                except Exception as e:
                    erreurs.append(f"Ligne {i} : {str(e)}")

            msg = f"{ajoutes} livre(s) importé(s), {ignores} ignoré(s)."
            if erreurs:
                msg += f" Erreurs : {'; '.join(erreurs)}"
            messages.success(request, msg)

        except FichierExcelInvalideException as e:
            messages.error(request, str(e))

    return redirect('/gestion-livres/')


# ══════════════════════════════════════════════════════════════
#  GESTION MEMBRES (Bibliothécaire)
# ══════════════════════════════════════════════════════════════
def gestion_membres(request):
    if 'user_id' not in request.session:
        return redirect('/login/')

    if request.method == 'POST':
        nom       = request.POST.get('nom', '').strip()
        prenom    = request.POST.get('prenom', '').strip()
        email     = request.POST.get('email', '').strip()
        telephone = request.POST.get('telephone', '').strip()
        login     = request.POST.get('login', '').strip()
        password  = request.POST.get('password', '').strip()

        try:
            if not nom or not email or not login or not password:
                raise ValueError("Nom, email, login et mot de passe sont obligatoires.")
            if Etudiant.objects.filter(email__iexact=email).exists():
                raise DoublonException(f"Un membre avec l'email '{email}' existe déjà.")
            if Utilisateur.objects.filter(login__iexact=login).exists():
                raise DoublonException(f"Le login '{login}' est déjà utilisé.")

            user = Utilisateur.objects.create(
                login=login,
                mot_de_passe=password,
                role='etudiant',
                actif=True
            )
            Etudiant.objects.create(
                utilisateur=user,
                nom=nom, prenom=prenom,
                email=email, telephone=telephone
            )
            messages.success(request, f"Membre '{nom} {prenom}' ajouté avec login '{login}'.")

        except (DoublonException, ValueError) as e:
            messages.error(request, str(e))

    # Suppression
    suppr_id = request.GET.get('suppr')
    if suppr_id:
        try:
            membre = Etudiant.objects.get(id=suppr_id)
            if Emprunt.objects.filter(etudiant=membre, rendu=False).exists():
                raise EmpruntEnCoursException(
                    f"Impossible — '{membre.nom}' a des emprunts en cours."
                )
            if membre.utilisateur:
                membre.utilisateur.delete()
            else:
                membre.delete()
            messages.success(request, "Membre supprimé.")
        except Etudiant.DoesNotExist:
            messages.error(request, "Membre introuvable.")
        except EmpruntEnCoursException as e:
            messages.error(request, str(e))

    return render(request, 'gestion_membres.html', {
        'membres': Etudiant.objects.all().order_by('nom')
    })


# ── Import Excel Étudiants ─────────────────────────────────────
def import_etudiants_excel(request):
    if 'user_id' not in request.session:
        return redirect('/login/')

    if request.method == 'POST':
        fichier = request.FILES.get('fichier_excel')

        try:
            if not fichier:
                raise FichierExcelInvalideException("Aucun fichier sélectionné.")
            if not fichier.name.endswith(('.xlsx', '.xls')):
                raise FichierExcelInvalideException("Le fichier doit être au format Excel (.xlsx).")

            wb      = openpyxl.load_workbook(fichier)
            ws      = wb.active
            headers = [str(cell.value).strip().lower() if cell.value else ''
                       for cell in ws[1]]

            colonnes_requises = ['nom', 'email', 'login', 'password']
            for col in colonnes_requises:
                if col not in headers:
                    raise FichierExcelInvalideException(
                        f"Colonne '{col}' manquante. Colonnes requises : nom, prenom, email, telephone, login, password"
                    )

            ajoutes = 0
            ignores = 0
            erreurs = []

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    data      = dict(zip(headers, row))
                    nom       = str(data.get('nom', '') or '').strip()
                    prenom    = str(data.get('prenom', '') or '').strip()
                    email     = str(data.get('email', '') or '').strip()
                    telephone = str(data.get('telephone', '') or '').strip()
                    login     = str(data.get('login', '') or '').strip()
                    password  = str(data.get('password', '') or '').strip()

                    if not nom or not email or not login or not password:
                        ignores += 1
                        continue

                    if Etudiant.objects.filter(email__iexact=email).exists() or \
                       Utilisateur.objects.filter(login__iexact=login).exists():
                        ignores += 1
                        continue

                    user = Utilisateur.objects.create(
                        login=login,
                        mot_de_passe=password,
                        role='etudiant',
                        actif=True
                    )
                    Etudiant.objects.create(
                        utilisateur=user,
                        nom=nom, prenom=prenom,
                        email=email, telephone=telephone
                    )
                    ajoutes += 1

                except Exception as e:
                    erreurs.append(f"Ligne {i} : {str(e)}")

            msg = f"{ajoutes} étudiant(s) importé(s), {ignores} ignoré(s)."
            if erreurs:
                msg += f" Erreurs : {'; '.join(erreurs)}"
            messages.success(request, msg)

        except FichierExcelInvalideException as e:
            messages.error(request, str(e))

    return redirect('/gestion-membres/')


# ══════════════════════════════════════════════════════════════
#  GESTION EMPRUNTS (Bibliothécaire)
# ══════════════════════════════════════════════════════════════
def gestion_emprunts(request):
    if 'user_id' not in request.session:
        return redirect('/login/')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'emprunt':
            try:
                etudiant_id = request.POST.get('etudiant_id')
                livre_id    = request.POST.get('livre_id')
                date_retour = request.POST.get('date_retour')

                try:
                    etudiant = Etudiant.objects.get(id=etudiant_id)
                except Etudiant.DoesNotExist:
                    raise MembreIntrouvableException("Étudiant introuvable.")

                try:
                    livre = Livre.objects.get(id=livre_id)
                except Livre.DoesNotExist:
                    raise LivreIntrouvableException("Livre introuvable.")

                if not livre.disponible:
                    raise LivreIndisponibleException(
                        f"'{livre.titre}' est déjà emprunté."
                    )

                Emprunt.objects.create(
                    etudiant=etudiant,
                    livre=livre,
                    date_retour_prevue=date_retour
                )
                livre.disponible = False
                livre.save()
                messages.success(
                    request,
                    f"Emprunt enregistré — '{livre.titre}' pour {etudiant.nom}."
                )

            except (LivreIndisponibleException,
                    LivreIntrouvableException,
                    MembreIntrouvableException) as e:
                messages.error(request, str(e))

        elif action == 'retour':
            try:
                emprunt_id      = request.POST.get('emprunt_id')
                date_retour_eff = request.POST.get('date_retour_effectif')

                try:
                    emprunt = Emprunt.objects.get(id=emprunt_id, rendu=False)
                except Emprunt.DoesNotExist:
                    raise EmpruntIntrouvableException(
                        "Emprunt introuvable ou déjà rendu."
                    )

                dt_eff = datetime.strptime(date_retour_eff, '%Y-%m-%d').date()
                retard = max(0, (dt_eff - emprunt.date_retour_prevue).days)

                emprunt.date_retour_effective = dt_eff
                emprunt.retard = retard
                emprunt.rendu  = True
                emprunt.save()

                emprunt.livre.disponible = True
                emprunt.livre.save()

                if retard > 0:
                    messages.warning(
                        request,
                        f"Retour enregistré avec {retard} jour(s) de retard."
                    )
                else:
                    messages.success(request, "Retour enregistré à temps !")

            except EmpruntIntrouvableException as e:
                messages.error(request, str(e))

    # Emprunts en retard
    aujourd_hui    = date.today()
    retards        = Emprunt.objects.filter(
        rendu=False,
        date_retour_prevue__lt=aujourd_hui
    )
    nb_retards     = retards.count()

    return render(request, 'gestion_emprunts.html', {
        'emprunts':   Emprunt.objects.filter(rendu=False).order_by('date_retour_prevue'),
        'etudiants':  Etudiant.objects.all(),
        'livres':     Livre.objects.filter(disponible=True),
        'retards':    retards,
        'nb_retards': nb_retards,
    })


# ══════════════════════════════════════════════════════════════
#  NOTIFICATIONS RETARD
# ══════════════════════════════════════════════════════════════\

def notifications_retard(request):
    if 'user_id' not in request.session:
        return redirect('/login/')

    aujourd_hui = date.today()
    emprunt_id = request.GET.get('emprunt')

    if emprunt_id:
        try:
            emprunt = Emprunt.objects.get(id=emprunt_id, rendu=False, date_retour_prevue__lt=aujourd_hui)
            emprunt.notif_envoyee = True
            emprunt.save()
            messages.success(request, f"Notification envoyée à {emprunt.etudiant.nom} {emprunt.etudiant.prenom} pour '{emprunt.livre.titre}'.")
        except Emprunt.DoesNotExist:
            messages.error(request, "Emprunt introuvable ou déjà rendu.")
    else:
        retards = Emprunt.objects.filter(
            rendu=False,
            date_retour_prevue__lt=aujourd_hui,
            notif_envoyee=False
        )
        nb = 0
        for emprunt in retards:
            emprunt.notif_envoyee = True
            emprunt.save()
            nb += 1
        messages.success(request, f"{nb} notification(s) de retard envoyée(s).")

    return redirect('/notifications/')


# ══════════════════════════════════════════════════════════════
#  GESTION COMPTES (Administrateur)
# ══════════════════════════════════════════════════════════════
def import_etudiants(request):
    if request.method == "POST":
        # traitement du fichier Excel
        pass
    return render(request, 'import_etudiants.html')
def gestion_comptes(request):
    if 'user_id' not in request.session:
        return redirect('/login/')

    current_user_id = request.session.get('user_id')

    if request.method == 'POST':
        compte_id        = request.POST.get('compte_id')
        nouveau_login    = request.POST.get('nouveau_login', '').strip()
        nouveau_role     = request.POST.get('nouveau_role', '').strip()
        nouveau_password = request.POST.get('nouveau_password', '').strip()
        actif            = request.POST.get('actif') == 'on'

        try:
            compte = Utilisateur.objects.get(id=compte_id)
            is_self = int(compte_id) == current_user_id

            # Vérifier doublon login
            if Utilisateur.objects.filter(login=nouveau_login)\
                                .exclude(id=compte_id).exists():
                raise DoublonException(f"Le login '{nouveau_login}' est déjà utilisé.")
            
            compte.login = nouveau_login

            if is_self:
                if not actif:
                    messages.error(request, "Vous ne pouvez pas désactiver votre propre compte administrateur.")
                compte.role = 'administrateur'
                compte.actif = True
            else:
                compte.role = nouveau_role
                compte.actif = actif

            if nouveau_password:
                compte.mot_de_passe = nouveau_password

            compte.save()
            messages.success(request, "Compte modifié avec succès.")
        except Utilisateur.DoesNotExist:
            messages.error(request, "Compte introuvable.")
        except DoublonException as e:
            messages.error(request, str(e))

    # Suppression
    suppr_id = request.GET.get('suppr')
    if suppr_id:
        try:
            if int(suppr_id) == current_user_id:
                messages.error(request, "Vous ne pouvez pas supprimer votre propre compte.")
            else:
                Utilisateur.objects.get(id=suppr_id).delete()
                messages.success(request, "Compte supprimé.")
        except Utilisateur.DoesNotExist:
            messages.error(request, "Compte introuvable.")

    # Activer / Désactiver
    toggle_id = request.GET.get('toggle')
    if toggle_id:
        try:
            if int(toggle_id) == current_user_id:
                messages.error(request, "Vous ne pouvez pas désactiver votre propre compte administrateur.")
            else:
                compte = Utilisateur.objects.get(id=toggle_id)
                compte.actif = not compte.actif
                compte.save()
                statut = "activé" if compte.actif else "désactivé"
                messages.success(request, f"Compte {statut}.")
        except Utilisateur.DoesNotExist:
            messages.error(request, "Compte introuvable.")

    # Débloquer un compte
    debloquer_id = request.GET.get('debloquer')
    if debloquer_id:
        try:
            compte = Utilisateur.objects.get(id=debloquer_id)
            login = compte.login
            if login in tentatives:
                tentatives[login] = 0

            compte.actif = True
            compte.save()
            messages.success(request, f"Compte de {login} débloqué.")
        except Utilisateur.DoesNotExist:
            messages.error(request, "Compte introuvable.")

    role_filtre = request.GET.get('role', '').strip()
    comptes = Utilisateur.objects.all()
    if role_filtre:
        comptes = comptes.filter(role=role_filtre)
    comptes = comptes.order_by('role', 'login')

    for c in comptes:
        c.bloque = tentatives.get(c.login, 0) >= 3

    return render(request, 'gestion_comptes.html', {
        'comptes': comptes,
        'role_filtre': role_filtre,
    })
def retards_page(request):
    if 'user_id' not in request.session:
        return redirect('/login/')

    aujourd_hui = date.today()
    retards = Emprunt.objects.filter(
        rendu=False,
        date_retour_prevue__lt=aujourd_hui
    )
    nb_retards = retards.count()

    return render(request, 'notifications.html', {
        'retards': retards,
        'nb_retards': nb_retards,
    })

# ══════════════════════════════════════════════════════════════
#  RAPPORT (Administrateur)
# ══════════════════════════════════════════════════════════════
def rapport(request):
    if 'user_id' not in request.session:
        return redirect('/login/')

    if request.method == 'POST':
        type_rapport = request.POST.get('type_rapport')
        contenu      = ""

        if type_rapport == 'emprunts':
            emprunts = Emprunt.objects.all()
            contenu  = f"Total emprunts : {emprunts.count()}\n"
            contenu += f"En cours : {emprunts.filter(rendu=False).count()}\n"
            contenu += f"Rendus   : {emprunts.filter(rendu=True).count()}"

        elif type_rapport == 'retards':
            retards = Emprunt.objects.filter(
                rendu=False,
                date_retour_prevue__lt=date.today()
            )
            contenu = f"Emprunts en retard : {retards.count()}\n"
            for e in retards:
                jours = (date.today() - e.date_retour_prevue).days
                contenu += f"- {e.etudiant.nom} → {e.livre.titre} ({jours}j de retard)\n"

        elif type_rapport == 'membres':
            contenu  = f"Total membres : {Etudiant.objects.count()}\n"
            contenu += f"Comptes actifs : {Utilisateur.objects.filter(actif=True).count()}\n"
            contenu += f"Comptes inactifs : {Utilisateur.objects.filter(actif=False).count()}"

        Rapport.objects.create(type=type_rapport, contenu=contenu)
        messages.success(request, "Rapport généré avec succès !")

    rapports = Rapport.objects.all().order_by('-date_generation')
    return render(request, 'rapport.html', {'rapports': rapports})
def import_excel_page(request):
    if 'user_id' not in request.session:
        return redirect('/login/')
    return render(request, 'import_excel.html')


# ── Rapport PDF ────────────────────────────────────────────────
def rapport_pdf(request, rapport_id):
    if 'user_id' not in request.session:
        return redirect('/login/')

    try:
        r = Rapport.objects.get(id=rapport_id)
    except Rapport.DoesNotExist:
        messages.error(request, "Rapport introuvable.")
        return redirect('/rapport/')

    buffer   = BytesIO()
    p        = canvas.Canvas(buffer, pagesize=A4)
    largeur, hauteur = A4

    # En-tête
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, hauteur - 60, "Bibliothèque — Rapport")
    p.setFont("Helvetica", 12)
    p.drawString(50, hauteur - 90,
                 f"Type : {r.get_type_display()}")
    p.drawString(50, hauteur - 110,
                 f"Généré le : {r.date_generation.strftime('%d/%m/%Y %H:%M')}")

    # Ligne séparatrice
    p.line(50, hauteur - 125, largeur - 50, hauteur - 125)

    # Contenu
    p.setFont("Helvetica", 11)
    y = hauteur - 150
    for ligne in r.contenu.split('\n'):
        if y < 60:
            p.showPage()
            y = hauteur - 60
        p.drawString(50, y, ligne)
        y -= 20

    p.showPage()
    p.save()

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = \
        f'attachment; filename="rapport_{r.type}_{r.id}.pdf"'
    return response